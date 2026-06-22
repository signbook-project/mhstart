"""Generate Map Listings (startups/enablers) bulk-upload Excel template."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

NAVY = '0A2463'
SAFFRON = 'FF6B35'
WHITE = 'FFFFFF'

wb = Workbook()

# ============================================================
# SHEET 1: Instructions
# ============================================================
ws_info = wb.active
ws_info.title = 'Instructions'
ws_info.sheet_view.showGridLines = False

title_font = Font(name='Arial', size=18, bold=True, color=NAVY)
body_font = Font(name='Arial', size=11, color='1A1208')
bold_body = Font(name='Arial', size=11, bold=True, color='1A1208')

ws_info['B2'] = 'MHStart — Map Listings Bulk Upload Template'
ws_info['B2'].font = title_font
ws_info.merge_cells('B2:H2')

ws_info['B4'] = 'How to use this template'
ws_info['B4'].font = Font(name='Arial', size=13, bold=True, color=SAFFRON)
ws_info.merge_cells('B4:H4')

instructions = [
    '1. Go to the "Listings Data" sheet (tab at the bottom).',
    '2. Fill in one row per startup, incubator, VC, accelerator or other ecosystem entity.',
    '3. Name, Type, Contact Name, Email and City are required fields.',
    '4. Type must be one of the values in the dropdown: startup, incubator, vc, accelerator, angel, government, corporate, other.',
    '5. Latitude & Longitude are needed to show a pin on the map — look up your address on Google Maps and right-click to copy coordinates.',
    '6. Sectors should be comma-separated, e.g.  FinTech, SaaS, AI/ML',
    '7. Status must be one of: pending, active, paused, rejected (defaults to "pending" if blank — use Approve in admin to make it live).',
    '8. Save the file and upload it from Admin → Map Listings → "Bulk Upload" button.',
    '9. You will see a summary of how many rows were imported successfully and any errors.',
]
r = 6
for line in instructions:
    ws_info[f'B{r}'] = line
    ws_info[f'B{r}'].font = body_font
    ws_info.merge_cells(f'B{r}:H{r}')
    r += 1

r += 1
ws_info[f'B{r}'] = 'Column Reference'
ws_info[f'B{r}'].font = Font(name='Arial', size=13, bold=True, color=SAFFRON)
ws_info.merge_cells(f'B{r}:H{r}')
r += 2

col_ref = [
    ('Name*', 'Organization / startup name (required)'),
    ('Type*', 'startup, incubator, vc, accelerator, angel, government, corporate, other (required)'),
    ('Tagline', 'One-line description'),
    ('Description', 'Longer description shown on the map detail panel'),
    ('Logo URL', 'Direct link to a logo image (optional)'),
    ('Website', 'Organization website URL'),
    ('Contact Name*', 'Primary contact person (required)'),
    ('Email*', 'Contact email (required)'),
    ('Phone', 'Contact phone number'),
    ('Address', 'Full street address'),
    ('City*', 'City name (required)'),
    ('District', 'Maharashtra district'),
    ('Latitude', 'Map pin latitude, e.g. 19.0760'),
    ('Longitude', 'Map pin longitude, e.g. 72.8777'),
    ('Sectors', 'Comma-separated industry sectors, e.g. FinTech, AI/ML'),
    ('Stage', 'For startups: Idea, Pre-Seed, Seed, Series A, Series B+, Growth, Profitable'),
    ('Team Size', '1-5, 6-15, 16-50, 51-200, 200+'),
    ('Founded Year', 'e.g. 2021'),
    ('LinkedIn', 'LinkedIn page URL'),
    ('Twitter', 'Twitter/X page URL'),
    ('Instagram', 'Instagram page URL'),
    ('Status', 'pending / active / paused / rejected — defaults to "pending"'),
]
for label, desc in col_ref:
    ws_info[f'B{r}'] = label
    ws_info[f'B{r}'].font = bold_body
    ws_info[f'C{r}'] = desc
    ws_info[f'C{r}'].font = body_font
    ws_info.merge_cells(f'C{r}:H{r}')
    r += 1

ws_info.column_dimensions['A'].width = 2
ws_info.column_dimensions['B'].width = 20
for c in 'CDEFGH':
    ws_info.column_dimensions[c].width = 14

# ============================================================
# SHEET 2: Listings Data
# ============================================================
ws = wb.create_sheet('Listings Data')
headers = [
    'Name', 'Type', 'Tagline', 'Description', 'Logo URL', 'Website',
    'Contact Name', 'Email', 'Phone', 'Address', 'City', 'District',
    'Latitude', 'Longitude', 'Sectors', 'Stage', 'Team Size', 'Founded Year',
    'LinkedIn', 'Twitter', 'Instagram', 'Status',
]

header_font = Font(name='Arial', size=12, bold=True, color=WHITE)
header_fill = PatternFill('solid', start_color=NAVY)
thin = Side(style='thin', color='D0D0D0')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for i, h in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=i, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border
ws.row_dimensions[1].height = 32

examples = [
    ['AgroNext Technologies', 'startup', 'Smart irrigation for Indian farms', 'AgroNext builds IoT-based irrigation systems helping farmers save water and increase yield.',
     '', 'https://agronext.in', 'Priya Sharma', 'priya@agronext.in', '+91 98765 11111', 'FC Road, Shivajinagar', 'Pune', 'Pune',
     18.5246, 73.8786, 'AgriTech, IoT', 'Seed', '6-15', 2022, 'https://linkedin.com/company/agronext', '', '', 'active'],
    ['Maharashtra Innovation Hub', 'incubator', 'Govt-backed incubator for deep tech startups', 'A state-supported incubation center providing funding, mentorship and lab access.',
     '', 'https://mih.gov.in', 'Anil Joshi', 'contact@mih.gov.in', '+91 98765 22222', 'Hinjewadi Phase 2', 'Pune', 'Pune',
     18.5912, 73.7389, 'DeepTech, HealthTech', '', '', 2018, 'https://linkedin.com/company/mih', '', '', 'active'],
    ['Sahyadri Capital Partners', 'vc', 'Early-stage VC fund for Maharashtra startups', 'Invests ₹50L-5Cr in pre-seed and seed stage startups across Maharashtra.',
     '', 'https://sahyadricapital.com', 'Neha Rane', 'neha@sahyadricapital.com', '+91 98765 33333', 'BKC', 'Mumbai', 'Mumbai',
     19.0654, 72.8682, 'FinTech, SaaS, D2C', '', '', 2019, '', 'https://twitter.com/sahyadricap', '', 'active'],
]
for row_idx, row in enumerate(examples, start=2):
    for col_idx, val in enumerate(row, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.font = Font(name='Arial', size=10, color='5A5048')
        cell.alignment = Alignment(vertical='top', wrap_text=True)
        cell.border = border

for row_idx in range(5, 100):
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.border = border
        cell.alignment = Alignment(vertical='top', wrap_text=True)

widths = [24, 14, 28, 38, 24, 24, 18, 24, 16, 26, 14, 14, 12, 12, 24, 14, 12, 14, 24, 22, 22, 12]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

dv_type = DataValidation(type='list', formula1='"startup,incubator,vc,accelerator,angel,government,corporate,other"', allow_blank=False, showDropDown=False)
dv_type.error = 'Please select a valid type from the dropdown'
dv_type.errorTitle = 'Invalid Type'
ws.add_data_validation(dv_type)
dv_type.add('B2:B500')

dv_status = DataValidation(type='list', formula1='"pending,active,paused,rejected"', allow_blank=True, showDropDown=False)
ws.add_data_validation(dv_status)
dv_status.add('V2:V500')

dv_stage = DataValidation(type='list', formula1='"Idea,Pre-Seed,Seed,Series A,Series B+,Growth,Profitable"', allow_blank=True, showDropDown=False)
ws.add_data_validation(dv_stage)
dv_stage.add('P2:P500')

dv_team = DataValidation(type='list', formula1='"1-5,6-15,16-50,51-200,200+"', allow_blank=True, showDropDown=False)
ws.add_data_validation(dv_team)
dv_team.add('Q2:Q500')

ws.freeze_panes = 'A2'
ws.sheet_view.showGridLines = False

wb.save('/home/claude/mhstart/templates/map-listings-bulk-upload-template.xlsx')
print('Listings template saved.')
