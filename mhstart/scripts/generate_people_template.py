"""Generate People (team/advisors/partners) bulk-upload Excel template."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

NAVY = '0A2463'
SAFFRON = 'FF6B35'
WHITE = 'FFFFFF'

wb = Workbook()

ws_info = wb.active
ws_info.title = 'Instructions'
ws_info.sheet_view.showGridLines = False

title_font = Font(name='Arial', size=18, bold=True, color=NAVY)
body_font = Font(name='Arial', size=11, color='1A1208')
bold_body = Font(name='Arial', size=11, bold=True, color='1A1208')

ws_info['B2'] = 'MHStart — People Bulk Upload Template'
ws_info['B2'].font = title_font
ws_info.merge_cells('B2:H2')

ws_info['B4'] = 'How to use this template'
ws_info['B4'].font = Font(name='Arial', size=13, bold=True, color=SAFFRON)
ws_info.merge_cells('B4:H4')

instructions = [
    '1. Go to the "People Data" sheet (tab at the bottom).',
    '2. Fill in one row per person — team member, founding member, advisor, supporter or partner.',
    '3. Name and Category are required fields.',
    '4. Category must be one of: team, founding_member, advisor, supported_by, partner (use the dropdown).',
    '5. Photo URL is optional — if left blank, a colored initial badge will be shown instead.',
    '6. Display Order controls the sort order within each category (lower numbers show first).',
    '7. Save the file and upload it from Admin → People → "Bulk Upload" button.',
]
r = 6
for line in instructions:
    ws_info[f'B{r}'] = line
    ws_info[f'B{r}'].font = body_font
    ws_info.merge_cells(f'B{r}:H{r}')
    r += 1

r += 1
ws_info[f'B{r}'] = 'Column Reference'
ws_info[f'B{r}'].font = Font(name='Arial', size=13, bold=True, color=SAFFRON)
ws_info.merge_cells(f'B{r}:H{r}')
r += 2

col_ref = [
    ('Name*', 'Full name (required)'),
    ('Role', 'Title, e.g. Co-Founder & CEO'),
    ('Organization', 'Company / institution name'),
    ('Category*', 'team, founding_member, advisor, supported_by, partner (required)'),
    ('Bio', 'Short biography'),
    ('Photo URL', 'Direct link to profile photo (optional)'),
    ('Email', 'Contact email'),
    ('LinkedIn', 'LinkedIn profile URL'),
    ('Twitter', 'Twitter/X profile URL'),
    ('Website', 'Personal or company website'),
    ('Display Order', 'Number controlling sort order, e.g. 1, 2, 3'),
    ('Visible (Yes/No)', 'Whether to show on the website — defaults to Yes'),
]
for label, desc in col_ref:
    ws_info[f'B{r}'] = label
    ws_info[f'B{r}'].font = bold_body
    ws_info[f'C{r}'] = desc
    ws_info[f'C{r}'].font = body_font
    ws_info.merge_cells(f'C{r}:H{r}')
    r += 1

ws_info.column_dimensions['A'].width = 2
ws_info.column_dimensions['B'].width = 20
for c in 'CDEFGH':
    ws_info.column_dimensions[c].width = 14

ws = wb.create_sheet('People Data')
headers = ['Name', 'Role', 'Organization', 'Category', 'Bio', 'Photo URL', 'Email', 'LinkedIn', 'Twitter', 'Website', 'Display Order', 'Visible (Yes/No)']

header_font = Font(name='Arial', size=12, bold=True, color=WHITE)
header_fill = PatternFill('solid', start_color=NAVY)
thin = Side(style='thin', color='D0D0D0')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for i, h in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=i, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border
ws.row_dimensions[1].height = 32

examples = [
    ['Dr. Rajesh Kumar', 'Co-Founder & CEO', 'MHStart', 'team', 'Serial entrepreneur with 15 years in the startup ecosystem.',
     '', 'rajesh@mhstart.com', 'https://linkedin.com/in/rajeshkumar', '', '', 1, 'Yes'],
    ['Anita Deshpande', 'Strategic Advisor', 'Sahyadri Capital', 'advisor', 'Former VC partner advising on fundraising strategy.',
     '', 'anita@sahyadricapital.com', 'https://linkedin.com/in/anitadeshpande', '', '', 1, 'Yes'],
]
for row_idx, row in enumerate(examples, start=2):
    for col_idx, val in enumerate(row, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.font = Font(name='Arial', size=10, color='5A5048')
        cell.alignment = Alignment(vertical='top', wrap_text=True)
        cell.border = border

for row_idx in range(4, 60):
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.border = border
        cell.alignment = Alignment(vertical='top', wrap_text=True)

widths = [22, 22, 22, 16, 40, 24, 24, 26, 22, 22, 12, 14]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

dv_cat = DataValidation(type='list', formula1='"team,founding_member,advisor,supported_by,partner"', allow_blank=False, showDropDown=False)
dv_cat.error = 'Please select a valid category from the dropdown'
dv_cat.errorTitle = 'Invalid Category'
ws.add_data_validation(dv_cat)
dv_cat.add('D2:D500')

dv_vis = DataValidation(type='list', formula1='"Yes,No"', allow_blank=True, showDropDown=False)
ws.add_data_validation(dv_vis)
dv_vis.add('L2:L500')

ws.freeze_panes = 'A2'
ws.sheet_view.showGridLines = False

wb.save('/home/claude/mhstart/templates/people-bulk-upload-template.xlsx')
print('People template saved.')
