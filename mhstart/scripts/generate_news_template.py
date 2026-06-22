"""Generate News bulk-upload Excel template."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

NAVY = '0A2463'
SAFFRON = 'FF6B35'
GRAY = 'F0EDE8'
WHITE = 'FFFFFF'

wb = Workbook()

# ============================================================
# SHEET 1: Instructions
# ============================================================
ws_info = wb.active
ws_info.title = 'Instructions'
ws_info.sheet_view.showGridLines = False

title_font = Font(name='Arial', size=18, bold=True, color=NAVY)
header_font = Font(name='Arial', size=12, bold=True, color=WHITE)
body_font = Font(name='Arial', size=11, color='1A1208')
bold_body = Font(name='Arial', size=11, bold=True, color='1A1208')

ws_info['B2'] = 'MHStart — News Bulk Upload Template'
ws_info['B2'].font = title_font
ws_info.merge_cells('B2:H2')

ws_info['B4'] = 'How to use this template'
ws_info['B4'].font = Font(name='Arial', size=13, bold=True, color=SAFFRON)
ws_info.merge_cells('B4:H4')

instructions = [
    '1. Go to the "News Data" sheet (tab at the bottom).',
    '2. Fill in one row per news article. Do NOT change the column headers in row 1.',
    '3. Title, Excerpt, Content and Author Name are required fields.',
    '4. Status must be one of: pending, published, rejected (use the dropdown).',
    '5. Tags should be comma-separated, e.g.  funding, fintech, pune',
    '6. Cover Image URL is optional — paste a direct image link (must start with http:// or https://).',
    '7. Content can include basic HTML tags like <p>, <b>, <i>, <a href="...">, <img src="...">.',
    '8. Save the file and upload it from Admin → News → "Bulk Upload" button.',
    '9. You will see a summary of how many rows were imported successfully and any errors.',
]
r = 6
for line in instructions:
    ws_info[f'B{r}'] = line
    ws_info[f'B{r}'].font = body_font
    ws_info.merge_cells(f'B{r}:H{r}')
    r += 1

r += 1
ws_info[f'B{r}'] = 'Column Reference'
ws_info[f'B{r}'].font = Font(name='Arial', size=13, bold=True, color=SAFFRON)
ws_info.merge_cells(f'B{r}:H{r}')
r += 2

col_ref = [
    ('Title*', 'The news headline (required)'),
    ('Excerpt*', 'Short 1-2 sentence summary shown in listing cards (required)'),
    ('Content*', 'Full article text/HTML (required)'),
    ('Cover Image URL', 'Direct link to a banner image (optional)'),
    ('Author Name*', 'Name of the person/org submitting the news (required)'),
    ('Author Email', 'Contact email of the author (optional)'),
    ('Tags', 'Comma-separated keywords, e.g. funding, ai, mumbai'),
    ('Status', 'pending / published / rejected — defaults to "pending" if blank'),
    ('Pinned (Yes/No)', 'Whether to feature this article at the top — defaults to No'),
]
for label, desc in col_ref:
    ws_info[f'B{r}'] = label
    ws_info[f'B{r}'].font = bold_body
    ws_info[f'C{r}'] = desc
    ws_info[f'C{r}'].font = body_font
    ws_info.merge_cells(f'C{r}:H{r}')
    r += 1

ws_info.column_dimensions['A'].width = 2
ws_info.column_dimensions['B'].width = 22
for c in 'CDEFGH':
    ws_info.column_dimensions[c].width = 14

# ============================================================
# SHEET 2: News Data
# ============================================================
ws = wb.create_sheet('News Data')
headers = ['Title', 'Excerpt', 'Content', 'Cover Image URL', 'Author Name', 'Author Email', 'Tags', 'Status', 'Pinned (Yes/No)']

header_fill = PatternFill('solid', start_color=NAVY)
thin = Side(style='thin', color='D0D0D0')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for i, h in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=i, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border
ws.row_dimensions[1].height = 32

# Example rows
examples = [
    ['Pune Startup Raises ₹5 Cr Seed Round', 'AgriTech startup AgroNext closes seed funding led by regional angels.',
     '<p>AgroNext, a Pune-based agritech startup, today announced it has raised ₹5 crore in a seed funding round...</p>',
     'https://example.com/images/agronext.jpg', 'Priya Sharma', 'priya@agronext.in', 'funding, agritech, pune', 'published', 'Yes'],
    ['MHStart Hosts Founders Meetup in Nagpur', 'Over 100 founders gathered for networking and panel discussions.',
     '<p>The Nagpur Founders Meetup brought together startups, investors and mentors for a day of networking...</p>',
     '', 'Rahul Deshmukh', 'rahul@mhstart.com', 'event, nagpur, networking', 'pending', 'No'],
]
for row_idx, row in enumerate(examples, start=2):
    for col_idx, val in enumerate(row, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.font = Font(name='Arial', size=10, color='5A5048')
        cell.alignment = Alignment(vertical='top', wrap_text=True)
        cell.border = border

# Empty rows for user data (style only)
for row_idx in range(4, 60):
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.border = border
        cell.alignment = Alignment(vertical='top', wrap_text=True)

# Column widths
widths = [30, 36, 46, 30, 18, 22, 24, 14, 14]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Data validation: Status dropdown
dv_status = DataValidation(type='list', formula1='"pending,published,rejected"', allow_blank=True, showDropDown=False)
dv_status.error = 'Please select: pending, published, or rejected'
dv_status.errorTitle = 'Invalid Status'
ws.add_data_validation(dv_status)
dv_status.add(f'H2:H500')

# Data validation: Pinned Yes/No
dv_pin = DataValidation(type='list', formula1='"Yes,No"', allow_blank=True, showDropDown=False)
ws.add_data_validation(dv_pin)
dv_pin.add('I2:I500')

ws.freeze_panes = 'A2'
ws.sheet_view.showGridLines = False

wb.save('/home/claude/mhstart/templates/news-bulk-upload-template.xlsx')
print('News template saved.')
